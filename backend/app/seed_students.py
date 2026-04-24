from openpyxl import load_workbook
import re
from pathlib import Path

from database import SessionLocal
from models.student import Student

ROSTER_FILE = Path(__file__).with_name("REGISTRO ACADEMICO ROSTER.xlsx")
SHEET_NAME = "ROSTER REGISTRO ACADEMICO"


def norm_code(value: str) -> str:
    return str(value).strip().upper().replace(" ", "")


def is_space(value: str) -> bool:
    return norm_code(value) == "SPACE"


def split_name(full_name: str):
    """
    Heurística simple:
    - si 1 token: first_name = full, last_name = ""
    - si 2 tokens: first_name = token1, last_name = token2
    - si >=3 tokens: last_name = últimos 2, first_name = resto
    """
    parts = [p for p in re.split(r"\s+", str(full_name).strip()) if p]
    if len(parts) == 0:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    if len(parts) == 2:
        return parts[0], parts[1]
    return " ".join(parts[:-2]), " ".join(parts[-2:])


def run():
    db = SessionLocal()

    if not ROSTER_FILE.exists():
        db.close()
        raise FileNotFoundError(f"No se encontró el roster: {ROSTER_FILE}")

    wb = load_workbook(ROSTER_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    # localizar fila de encabezado (donde aparece CODE en la columna B)
    header_row = None
    for r in range(1, 20):
        v = ws.cell(row=r, column=2).value
        if v and str(v).strip().upper() == "CODE":
            header_row = r
            break
    if not header_row:
        db.close()
        raise RuntimeError("No se encontró encabezado CODE en el roster.")

    created = 0
    updated = 0
    skipped = 0
    dup_codes_skipped = 0

    seen_codes = set()

    for r in range(header_row + 1, ws.max_row + 1):
        code_raw = ws.cell(row=r, column=2).value      # CODE (B)
        name_raw = ws.cell(row=r, column=3).value      # STUDENT (C)

        if code_raw is None and name_raw is None:
            continue

        code = norm_code(code_raw)
        full = str(name_raw).strip() if name_raw is not None else ""

        # ignora vacíos / SPACE (si quedan)
        if not code or not full or is_space(code) or is_space(full):
            continue

        # evita duplicados de CODE dentro del mismo archivo
        if code in seen_codes:
            dup_codes_skipped += 1
            continue
        seen_codes.add(code)

        first_name, last_name = split_name(full)

        existing = db.query(Student).filter(Student.local_code == code).first()
        if existing:
            changed = False
            if existing.first_name != first_name:
                existing.first_name = first_name
                changed = True
            if existing.last_name != last_name:
                existing.last_name = last_name
                changed = True

            # mined_id se deja opcional (None)
            if changed:
                updated += 1
            else:
                skipped += 1
        else:
            db.add(Student(
                local_code=code,
                mined_id=None,           # Opción A: opcional
                first_name=first_name,
                last_name=last_name,
                active=True
            ))
            created += 1

    db.commit()
    db.close()

    print(f"✅ seed_students | created={created} updated={updated} skipped={skipped} dup_codes_skipped={dup_codes_skipped}")


if __name__ == "__main__":
    run()
