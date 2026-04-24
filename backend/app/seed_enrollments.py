import os
import re
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError

from database import SessionLocal
from models.student import Student
from models.grade import Grade
from models.section import Section
from models.enrollment import Enrollment

ROSTER_FILE = Path(__file__).with_name("REGISTRO ACADEMICO ROSTER.xlsx")
SHEET_NAME = "ROSTER REGISTRO ACADEMICO"

# ✅ Año variable por entorno
ACADEMIC_YEAR = int(os.getenv("ACADEMIC_YEAR", "2025"))

DEFAULT_SECTION_CODE = "A"


def norm(value) -> str:
    return str(value).strip().upper().replace(" ", "")


def is_space(value) -> bool:
    return norm(value) == "SPACE"


def roster_grade_to_internal(grade_raw: str) -> str:
    """
    Convierte: 1ST, 2ND, 3RD, 10TH, 11TH -> '1°'...'11°'
    """
    g = norm(grade_raw)
    if not g or g == "SPACE":
        return ""
    m = re.match(r"^(\d{1,2})", g)
    if not m:
        return ""
    n = int(m.group(1))
    if n < 1 or n > 11:
        return ""
    return f"{n}°"


def run():
    db = SessionLocal()

    if not ROSTER_FILE.exists():
        db.close()
        raise FileNotFoundError(f"No se encontró el roster: {ROSTER_FILE}")

    wb = load_workbook(ROSTER_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    # localizar encabezado
    header_row = None
    for r in range(1, 20):
        v = ws.cell(row=r, column=2).value
        if v and str(v).strip().upper() == "CODE":
            header_row = r
            break
    if not header_row:
        db.close()
        raise RuntimeError("No se encontró encabezado CODE en el roster.")

    created = 0
    updated = 0
    skipped = 0
    missing_student = 0
    missing_grade = 0
    missing_section = 0
    dup_in_roster = 0

    seen_keys = set()

    for r in range(header_row + 1, ws.max_row + 1):
        code_raw = ws.cell(row=r, column=2).value   # CODE
        name_raw = ws.cell(row=r, column=3).value   # STUDENT
        grade_raw = ws.cell(row=r, column=5).value  # GRADE

        if code_raw is None and name_raw is None and grade_raw is None:
            continue

        code = norm(code_raw)
        full = str(name_raw).strip() if name_raw is not None else ""
        grade_name = roster_grade_to_internal(grade_raw)

        if not code or not full or is_space(code) or is_space(full):
            continue
        if not grade_name:
            continue

        key = (code, grade_name, ACADEMIC_YEAR)
        if key in seen_keys:
            dup_in_roster += 1
            continue
        seen_keys.add(key)

        student = db.query(Student).filter(Student.local_code == code).first()
        if not student:
            missing_student += 1
            continue

        grade = db.query(Grade).filter(Grade.name == grade_name).first()
        if not grade:
            missing_grade += 1
            continue

        section = db.query(Section).filter(
            Section.grade_id == grade.id,
            Section.academic_year == ACADEMIC_YEAR,
            Section.code == DEFAULT_SECTION_CODE
        ).first()
        if not section:
            missing_section += 1
            continue

        # ✅ IDEMPOTENTE: si existe matrícula, no insertar; solo sincronizar section_id
        existing = db.query(Enrollment).filter(
            Enrollment.student_id == student.id,
            Enrollment.grade_id == grade.id,
            Enrollment.academic_year == ACADEMIC_YEAR
        ).first()

        if existing:
            if hasattr(existing, "section_id") and existing.section_id != section.id:
                existing.section_id = section.id
                updated += 1
            else:
                skipped += 1
            continue

        enrollment = Enrollment(
            student_id=student.id,
            grade_id=grade.id,
            academic_year=ACADEMIC_YEAR
        )
        if hasattr(enrollment, "section_id"):
            enrollment.section_id = section.id

        db.add(enrollment)
        created += 1

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        print("⚠️ IntegrityError: hubo duplicados por UNIQUE constraint. Se omitieron.")

    db.close()

    print(
        "✅ seed_enrollments (idempotente) | "
        f"year={ACADEMIC_YEAR} created={created} updated={updated} skipped={skipped} "
        f"missing_student={missing_student} missing_grade={missing_grade} missing_section={missing_section} "
        f"dup_in_roster_skipped={dup_in_roster}"
    )


if __name__ == "__main__":
    run()
