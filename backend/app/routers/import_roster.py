import os
import re
import uuid
from pathlib import Path
from typing import Optional, Dict, Any, List

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from openpyxl import load_workbook

from app.database import get_db
from security import get_current_user, require_roles

from models.student import Student
from models.grade import Grade
from models.section import Section
from models.enrollment import Enrollment


router = APIRouter(prefix="/import", tags=["Import"])

ALLOWED = ("ADMIN", "COORDINATION")
DEFAULT_SECTION_CODE = "A"


# =====================================================
# Helpers
# =====================================================

def norm_code(v: str) -> str:
    return str(v).strip().upper().replace(" ", "")


def is_blank_or_space(v: str) -> bool:
    s = norm_code(v)
    return (not s) or s == "SPACE" or s == "0"


def split_name(full: str):
    parts = [p for p in re.split(r"\s+", str(full).strip()) if p]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def roster_grade_to_internal(grade_raw: str) -> str:
    g = norm_code(grade_raw)
    m = re.match(r"^(\d{1,2})", g)
    if not m:
        return ""
    return f"{int(m.group(1))}°"


def ensure_section_a(db: Session, grade: Grade, academic_year: int) -> Section:
    section = db.query(Section).filter(
        Section.grade_id == grade.id,
        Section.academic_year == academic_year,
        Section.code == DEFAULT_SECTION_CODE
    ).first()

    if not section:
        section = Section(
            grade_id=grade.id,
            academic_year=academic_year,
            code=DEFAULT_SECTION_CODE,
            name=f"{grade.name}{DEFAULT_SECTION_CODE}"
        )
        db.add(section)
        db.flush()

    return section


# =====================================================
# IMPORT ROSTER XLSX
# =====================================================

@router.post("/roster-xlsx")
async def import_roster_xlsx(
    file: UploadFile = File(...),
    academic_year: Optional[int] = None,
    overwrite_names: bool = True,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    require_roles(*ALLOWED)(user)

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Debe subir un archivo XLSX")

    year = academic_year or int(os.getenv("ACADEMIC_YEAR", "2025"))

    content = await file.read()
    tmp_path = Path(f"._tmp_roster_{uuid.uuid4().hex}.xlsx")
    tmp_path.write_bytes(content)

    warnings: List[Dict[str, Any]] = []
    students_created = students_updated = students_skipped = 0
    enrollments_created = 0

    try:
        wb = load_workbook(tmp_path, data_only=True)

        sheet_name = "ROSTER REGISTRO ACADEMICO"
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Hoja ROSTER REGISTRO ACADEMICO no encontrada")

        ws = wb[sheet_name]

        header_row = None
        for r in range(1, 25):
            if str(ws.cell(row=r, column=2).value).strip().upper() == "CODE":
                header_row = r
                break

        if not header_row:
            raise HTTPException(status_code=400, detail="No se encontró encabezado CODE")

        for r in range(header_row + 1, ws.max_row + 1):
            code_raw = ws.cell(row=r, column=2).value
            name_raw = ws.cell(row=r, column=3).value
            grade_raw = ws.cell(row=r, column=5).value

            if code_raw is None and name_raw is None:
                continue

            code = norm_code(code_raw)
            full_name = str(name_raw).strip() if name_raw else ""

            if is_blank_or_space(code):
                warnings.append({
                    "row": r,
                    "message": "Fila ignorada: no corresponde a un estudiante válido."
                })
                students_skipped += 1
                continue

            grade_name = roster_grade_to_internal(grade_raw)
            if not grade_name:
                warnings.append({
                    "row": r,
                    "code": code,
                    "message": "Fila ignorada: grado no válido."
                })
                students_skipped += 1
                continue

            first_name, last_name = split_name(full_name)
            if not first_name:
                warnings.append({
                    "row": r,
                    "code": code,
                    "message": "Fila ignorada: nombre vacío."
                })
                students_skipped += 1
                continue

            student = db.query(Student).filter(Student.local_code == code).first()

            if student:
                changed = False
                if overwrite_names and student.first_name != first_name:
                    student.first_name = first_name
                    changed = True
                if overwrite_names and student.last_name != last_name:
                    student.last_name = last_name
                    changed = True
                students_updated += int(changed)
                students_skipped += int(not changed)
            else:
                student = Student(
                    local_code=code,
                    mined_id=None,
                    first_name=first_name,
                    last_name=last_name,
                    active=True
                )
                db.add(student)
                db.flush()
                students_created += 1

            grade = db.query(Grade).filter(Grade.name == grade_name).first()
            if not grade:
                warnings.append({
                    "row": r,
                    "code": code,
                    "message": f"Fila ignorada: grado {grade_name} no existe."
                })
                continue

            section = ensure_section_a(db, grade, year)

            exists = db.query(Enrollment).filter(
                Enrollment.student_id == student.id,
                Enrollment.grade_id == grade.id,
                Enrollment.academic_year == year
            ).first()

            if not exists:
                db.add(Enrollment(
                    student_id=student.id,
                    grade_id=grade.id,
                    section_id=section.id,
                    academic_year=year
                ))
                enrollments_created += 1

        db.commit()

        response = {
            "academic_year": year,
            "summary": {
                "students_created": students_created,
                "students_updated": students_updated,
                "students_skipped": students_skipped,
                "enrollments_created": enrollments_created
            },
            "warnings_count": len(warnings),
            "warnings": warnings[:20],
        }

        if len(warnings) > 20:
            response["note"] = "Se muestran solo las primeras 20 advertencias."

        return response

    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        tmp_path.unlink(missing_ok=True)
