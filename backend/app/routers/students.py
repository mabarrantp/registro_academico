from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import get_db
from app.models.student import Student
from app.schemas.student import StudentCreate, StudentImportRow
from app.services.student_code import generate_student_code

router = APIRouter(prefix="/students", tags=["Students"])


@router.post("")
def create_student(
    payload: StudentCreate,
    db: Session = Depends(get_db)
):
    student_code = generate_student_code(
        db=db,
        entry_year=payload.entry_year,
        entry_grade_id=payload.entry_grade_id
    )

    student = Student(
        student_code=student_code,
        first_name=payload.first_name,
        last_name=payload.last_name,
        entry_year=payload.entry_year,
        entry_grade_id=payload.entry_grade_id,
        current_grade_id=payload.entry_grade_id,
        active=True
    )

    db.add(student)
    db.commit()
    db.refresh(student)

    return student


@router.post("/import-xlsx")
def import_students_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser Excel (.xlsx)")

    wb = load_workbook(file.file)
    ws = wb.active

    expected_headers = ["first_name", "last_name", "entry_year", "entry_grade_id"]
    headers = [cell.value for cell in ws[1]]

    if headers != expected_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Encabezados inválidos. Se espera: {expected_headers}"
        )

    created = 0
    errors = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = StudentImportRow(
                first_name=row[0],
                last_name=row[1],
                entry_year=int(row[2]),
                entry_grade_id=int(row[3]),
            )

            student_code = generate_student_code(
                db=db,
                entry_year=data.entry_year,
                entry_grade_id=data.entry_grade_id
            )

            student = Student(
                student_code=student_code,
                first_name=data.first_name.strip(),
                last_name=data.last_name.strip(),
                entry_year=data.entry_year,
                entry_grade_id=data.entry_grade_id,
                current_grade_id=data.entry_grade_id,
                active=True
            )

            db.add(student)
            db.commit()
            created += 1

        except Exception as e:
            db.rollback()
            errors.append(f"Fila {idx}: {str(e)}")

    return {
        "created": created,
        "errors": errors
    }